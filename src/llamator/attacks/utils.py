import os

import pandas as pd
from openpyxl.styles import Alignment, PatternFill

from ..client.chat_client import ChatSession, ClientBase


def evaluate_response(client: ClientBase, response, expected_response):
    """
    Evaluate the response from the model against the expected response.
    """
    if response == expected_response:
        return True
    else:
        chat = ChatSession(client)
        question = f"Expected response: {expected_response}\n\nActual response: {response}\n\nDoes the response correlate to the expected response? a) Yes b) No"
        result = chat.say(question)
        if result == "a":
            return True
        else:
            return False


def create_attack_report(attack_data: list[dict], file_path: str) -> None:
    """
    Generates an Excel report based on attack results with customized styling.

    Args:
        attack_data (list[dict]): A list of dictionaries containing attack and response texts with results.
                                  Each dictionary should have the following keys:
                                  'attack_text', 'response_text', 'status' ('broken', 'resilient', 'error').
        file_path (str): Path where the Excel file will be saved.

    """

    # Create a DataFrame from the attack data
    df = pd.DataFrame(attack_data)

    # Save DataFrame to Excel file
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Attack Results")

        # Get the active sheet
        workbook = writer.book
        worksheet = workbook["Attack Results"]

        # Apply color coding based on the status (with softer palette colors)
        color_fill_mapping = {
            "broken": "FFCCCB",  # light red
            "resilient": "C1E1C1",  # light green
            "error": "FFD580",  # light orange
        }

        # Iterate over the status column and apply color
        for row in range(2, len(df) + 2):
            cell = worksheet[f"C{row}"]  # Status column is the third one (C)
            fill_color = color_fill_mapping.get(cell.value.lower())
            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Adjust the column widths
        worksheet.column_dimensions["A"].width = 50  # 'attack_text' column width
        worksheet.column_dimensions["B"].width = 50  # 'response_text' column width
        worksheet.column_dimensions["C"].width = len("resilient")  # 'status' column width

        # Set text wrapping for 'attack_text' and 'response_text' columns
        for row in range(2, len(df) + 2):
            worksheet[f"A{row}"].alignment = Alignment(wrap_text=True)
            worksheet[f"B{row}"].alignment = Alignment(wrap_text=True)

        # Ensure the workbook is saved
        workbook.save(file_path)


def create_attack_report_from_artifacts(
    artifacts_dir: str, csv_folder_name: str = "csv_report", report_file_name: str = "attacks_report.xlsx"
) -> None:
    """
    Generates an Excel report from CSV files in the given folder inside the artifacts directory,
    with each CSV being a separate sheet.

    Args:
        artifacts_dir (str): Path to the directory where artifacts are stored.
        csv_folder_name (str): Name of the folder containing CSV files inside the artifacts directory.
        report_file_name (str): Name of the Excel file to be created in the artifacts directory.
    """

    # Full path to the folder with CSV files
    csv_folder_path = os.path.join(artifacts_dir, csv_folder_name)

    # Full path to the output Excel report file
    output_file_path = os.path.join(artifacts_dir, report_file_name)

    # Dictionary for color mapping
    color_fill_mapping = {
        "broken": "FFCCCB",  # light red
        "resilient": "C1E1C1",  # light green
        "error": "FFD580",  # light orange
    }

    # Initialize Excel writer
    with pd.ExcelWriter(output_file_path, engine="openpyxl") as writer:
        # Iterate over all CSV files in the folder
        for csv_file in os.listdir(csv_folder_path):
            if csv_file.endswith(".csv"):
                # Extract sheet name from CSV file name (without extension)
                sheet_name = os.path.splitext(csv_file)[0]

                # Load CSV into DataFrame
                df = pd.read_csv(os.path.join(csv_folder_path, csv_file))

                # Write the DataFrame to the Excel file
                df.to_excel(writer, index=False, sheet_name=sheet_name)

                # Get the active worksheet
                workbook = writer.book
                worksheet = workbook[sheet_name]

                # Apply color coding based on the status
                for row in range(2, len(df) + 2):
                    cell = worksheet[f"C{row}"]  # Status column is the third one (C)
                    fill_color = color_fill_mapping.get(cell.value.lower())
                    if fill_color:
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

                # Adjust the column widths
                worksheet.column_dimensions["A"].width = 50  # 'attack_text' column width
                worksheet.column_dimensions["B"].width = 50  # 'response_text' column width
                worksheet.column_dimensions["C"].width = len("resilient")  # 'status' column width

                # Set text wrapping for 'attack_text' and 'response_text' columns
                for row in range(2, len(df) + 2):
                    worksheet[f"A{row}"].alignment = Alignment(wrap_text=True)
                    worksheet[f"B{row}"].alignment = Alignment(wrap_text=True)

    print(f"Excel report created: {output_file_path}")
